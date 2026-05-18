"""
FIORA IC — Geração da Planilha de Setup CFD v4.0
=================================================
Atualiza sobre a v3.0:
- Aba 2: saídas DN50 reduzidas para 3 (não 4)
- Aba 8 (Eletrodos): verificação de folga com Sep. 20% (270 mm)
- Aba 9 (Regiões A/B): atualizar (Sep. 95% sem DN50 lateral)
- Aba 10 (Reports): atualizar boundaries (3 DN50 + manifold + 6 outlets)
- Aba 11 (Cenários): manter — sem mudança
- NOVA Aba 12 (Folgas Críticas): verificações sistemáticas do CAD v5.11
- NOVA Aba 13 (Manifold de 4 ramais DN50): parâmetros + diagrama

Preserva 100% a estrutura visual.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SRC = "/home/user/Programacao/FIORA_IC_Setup_CFD_v3.xlsx"
OUT = "/home/user/Programacao/FIORA_IC_Setup_CFD_v4.xlsx"

C_TITLE     = "1F3864"
C_SUBTITLE  = "ED7D31"
C_INPUT_HD  = "0070C0"
C_CALC_HD   = "00B050"
C_NOTE_HD   = "7030A0"
C_INPUT_BG  = "DCE6F1"
C_CALC_BG   = "E2EFDA"
C_HIGHLIGHT = "FFF2CC"
C_FIXED     = "FCE4D6"

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFFFF", size=11)
FONT_BLUE       = Font(color="FF0000FF")
FONT_BOLD       = Font(bold=True)
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color):
    return PatternFill("solid", fgColor=color)


def set_cell(ws, coord, value, bg=None, font=None, border=False, align=None):
    c = ws[coord]
    c.value = value
    if bg is not None: c.fill = fill(bg)
    if font is not None: c.font = font
    if border: c.border = BORDER
    if align is not None: c.alignment = align
    return c


def write_row(ws, row, label, value, unit=None, note=None, value_bg=C_CALC_BG, value_font=None):
    set_cell(ws, f"A{row}", label)
    set_cell(ws, f"B{row}", value, bg=value_bg, font=value_font)
    if unit is not None: set_cell(ws, f"C{row}", unit)
    if note is not None: set_cell(ws, f"D{row}", note)


def section_header(ws, row, text, color=C_INPUT_HD, cols=("A", "B", "C", "D")):
    for col in cols:
        set_cell(ws, f"{col}{row}", text if col == cols[0] else None,
                 bg=color, font=FONT_WHITE_BOLD)
        ws[f"{col}{row}"].border = BORDER
    ws.merge_cells(f"{cols[0]}{row}:{cols[-1]}{row}")
    ws[f"{cols[0]}{row}"].alignment = Alignment(horizontal="left", vertical="center")


# ═══════════════════════════════════════════════════════
# CARREGAR e ATUALIZAR VERSÃO
# ═══════════════════════════════════════════════════════
wb = load_workbook(SRC)
ws1 = wb["1_Geometria"]
ws1["B1"].value = "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v4.0  (18/Mai/2026)"

# ═══════════════════════════════════════════════════════
# ABA 2 — atualizar saídas DN50 (4 → 3)
# ═══════════════════════════════════════════════════════
ws2 = wb["2_Vazoes_Velocidades"]
for r in range(1, ws2.max_row + 1):
    a = ws2.cell(row=r, column=1).value
    if a and "Saídas separadores" in str(a):
        ws2.cell(row=r, column=1).value = "Saídas separadores — 3×DN50  ★ v4.0"
        ws2.cell(row=r, column=3).value = "az=0°; z=2,92/5,84/8,76m  (Sep. 95% sai INTERNO)"
        ws2.cell(row=r, column=3).fill = fill(C_FIXED)
        break

# ═══════════════════════════════════════════════════════
# ABA 8 — Eletrodos: adicionar verificação de folga com Sep. 20%
# ═══════════════════════════════════════════════════════
ws8 = wb["8_Eletrodos"]
# Encontrar última linha usada e adicionar bloco novo no final
last_row = ws8.max_row + 2
section_header(ws8, last_row, "★ v4.0 — VERIFICAÇÃO DE FOLGA CRÍTICA: SEP. 20% × ELETRODOS",
               color=C_NOTE_HD)
r = last_row + 1
write_row(ws8, r, "Cota topo dos eletrodos (z_top)",
          "=B8+B6", "m", "= ELEC_Z_BOT + ELEC_H = 2,65 m"); r += 1
write_row(ws8, r, "Cota nominal Sep. 20% (= borda inferior das placas)",
          "='1_Geometria'!C19", "m", "20% × H = 2,92 m"); r += 1
write_row(ws8, r, "Queda vertical das placas do Sep. 20%",
          "=(0.5*'1_Geometria'!C11/2)*SIN(RADIANS('1_Geometria'!C22))", "m",
          "W/2 × sin(47,5°) = 0,385 m (N=2, W=D/2)"); r += 1
write_row(ws8, r, "★ FOLGA = (Sep. 20% base) − (eletrodo topo)",
          f"=B{r-2}-B{r-3}", "m",
          "Deve ser > 50 mm. Calc.: 2,92 − 2,65 = 270 mm ✅"); r += 1
ws8[f"B{r-1}"].fill = fill(C_FIXED)

# ═══════════════════════════════════════════════════════
# ABA 9 — atualizar para refletir Sep. 95% sem DN50 lateral
# ═══════════════════════════════════════════════════════
ws9 = wb["9_Regioes_Cinetica_PosProc"]
last_row = ws9.max_row + 2
section_header(ws9, last_row, "★ v4.0 — ATUALIZAÇÃO DAS SAÍDAS DOS SEPARADORES",
               color=C_NOTE_HD)
r = last_row + 1
set_cell(ws9, f"A{r}",
         "Sep. 20/40/60% — DN50 lateral na parede (3 saídas a 0°, multifásico)\n"
         "Sep. 95% — saída INTERNA para câmara de gás do domo (sem DN50 lateral)\n"
         "Isso significa que o gás coletado no Sep. 95% sobe diretamente para o domo\n"
         "e sai pelo Outlet_Biogas central. NÃO há furo na parede em z=13,87m.")
ws9.merge_cells(f"A{r}:E{r}")
ws9[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws9.row_dimensions[r].height = 75

# ═══════════════════════════════════════════════════════
# ABA 10 — atualizar Reports (sem Outlet_DN50_95)
# ═══════════════════════════════════════════════════════
ws10 = wb["10_Reports_STARCCM"]
# Procurar e atualizar Q_DN50_95 → tornar nota
for r in range(1, ws10.max_row + 1):
    a = ws10.cell(row=r, column=1).value
    if a and "Q_DN50_95" in str(a):
        ws10.cell(row=r, column=1).value = "Q_DN50_95 ★ REMOVIDO v4.0"
        ws10.cell(row=r, column=1).fill = fill(C_FIXED)
        ws10.cell(row=r, column=5).value = "Sep. 95% agora sai INTERNO p/ domo — não existe mais DN50 lateral"
        break

# ═══════════════════════════════════════════════════════
# ABA 12 — NOVA: Folgas Críticas Verificadas
# ═══════════════════════════════════════════════════════
ws12 = wb.create_sheet("12_Folgas_Criticas")
for cl, w in [("A", 38), ("B", 16), ("C", 16), ("D", 14), ("E", 38)]:
    ws12.column_dimensions[cl].width = w

set_cell(ws12, "B1", "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v4.0",
         bg=C_TITLE, font=FONT_WHITE_BOLD)
ws12.merge_cells("B1:E1")
set_cell(ws12, "B2", "ABA 12 — VERIFICAÇÕES DE FOLGA CRÍTICAS DO CAD v5.11 (★ NOVA v4.0)",
         bg=C_SUBTITLE, font=FONT_WHITE_BOLD)
ws12.merge_cells("B2:E2")

r = 4
section_header(ws12, r, "FOLGAS VERTICAIS", cols=("A", "B", "C", "D", "E"))
r += 1
headers = ["Par de Componentes", "Cota A [m]", "Cota B [m]", "Folga [mm]", "Status / Observação"]
for i, h in enumerate(headers):
    set_cell(ws12, f"{chr(65+i)}{r}", h, bg=C_INPUT_HD, font=FONT_WHITE_BOLD, border=True)
r += 1

folgas_vert = [
    ("Topo Eletrodos × Borda inferior Sep. 20%", 2.650, 2.920, "+270", "✅ Margem confortável — corrigido v5.10"),
    ("Pico Sep. 95% × Saída Efluente Lateral", 14.090, 14.300, "+210", "✅ OK"),
    ("Pico Sep. 95% × Topo do Cilindro", 14.090, 14.339, "+249", "✅ OK"),
    ("Saída Efluente × Topo do Cilindro", 14.300, 14.339, "+39", "⚠ Pequena — Vinícius queria 'bem no topo'"),
    ("Manifold × Topo dos Bocais", 0.175, 0.232, "+57", "✅ OK (verticalmente separados)"),
    ("Topo Difusores × Manifold", 0.025, 0.125, "+100", "✅ OK"),
    ("Topo Eletrodos × Pico Sep. 20%", 2.650, 3.305, "+655", "✅ Muito confortável"),
]
for f in folgas_vert:
    for i, v in enumerate(f):
        set_cell(ws12, f"{chr(65+i)}{r}", v, border=True)
        if i == 3 and "+" in str(v):
            ws12[f"{chr(65+i)}{r}"].font = FONT_BOLD
    r += 1

r += 1
section_header(ws12, r, "FOLGAS RADIAIS / AZIMUTAIS", cols=("A", "B", "C", "D", "E"))
r += 1
for i, h in enumerate(["Par de Componentes", "Posição A", "Posição B", "Folga [mm]", "Status / Observação"]):
    set_cell(ws12, f"{chr(65+i)}{r}", h, bg=C_INPUT_HD, font=FONT_WHITE_BOLD, border=True)
r += 1

folgas_rad = [
    ("Ramal az=45° × Eletrodo az=0° (linha y=x)", "az 45°", "az 0°", "+297", "✅ Distância mínima em XY"),
    ("Ramais × Parede do Reator (extremidade r=0,7m)", "r=0,7 m", "r=1,045 m", "+345", "✅ OK"),
    ("Eletrodo r=0,68 (borda ext.) × Parede", "r=0,94 m", "r=1,045 m", "+105", "✅ OK (memorial v1.0)"),
    ("Eletrodo r=0,68 (borda int.) × DN100 (r=0,05)", "r=0,42 m", "r=0,05 m", "+370", "✅ Muito OK"),
    ("Difusor externo × Parede", "r=0,80 m", "r=1,045 m", "+245", "✅ OK"),
    ("Difusor interno × DN100 central", "r=0,35 m", "r=0,05 m", "+300", "✅ OK"),
    ("DN100 retorno (raio 50mm) × Λs centrais (vale)", "r=0,05 m", "vale entre Λs", "+102", "✅ Passa no vale entre Λs centrais"),
]
for f in folgas_rad:
    for i, v in enumerate(f):
        set_cell(ws12, f"{chr(65+i)}{r}", v, border=True)
        if i == 3:
            ws12[f"{chr(65+i)}{r}"].font = FONT_BOLD
    r += 1

r += 1
section_header(ws12, r, "FOLGAS ANGULARES (Δθ)", cols=("A", "B", "C", "D", "E"))
r += 1
for i, h in enumerate(["Par de Componentes", "Az A [°]", "Az B [°]", "Δθ [°]", "Status / Observação"]):
    set_cell(ws12, f"{chr(65+i)}{r}", h, bg=C_INPUT_HD, font=FONT_WHITE_BOLD, border=True)
r += 1

folgas_ang = [
    ("Ramais × Eletrodos", "45/135/225/315", "0/180", "45°", "✅ Mínima Δθ = 45°"),
    ("Ramais × Bocais", "45/135/225/315", "0/60/120/180/...", "15°", "✅ Mínima Δθ = 15°"),
    ("Ramais × Difusores (setores)", "45/135/225/315", "30/90/150/210/270/330", "15°", "✅ Mínima Δθ = 15°"),
    ("Bocais × Eletrodos", "0/60/120/180/...", "0/180", "0°", "✅ Sobreposição mas em z diferentes"),
    ("Difusores × Bocais", "30/90/150/...", "0/60/120/...", "30°", "✅ Difusores ENTRE bocais"),
]
for f in folgas_ang:
    for i, v in enumerate(f):
        set_cell(ws12, f"{chr(65+i)}{r}", v, border=True)
    r += 1

r += 2
section_header(ws12, r, "AÇÃO RECOMENDADA", color=C_NOTE_HD, cols=("A", "B", "C", "D", "E"))
r += 1
set_cell(ws12, f"A{r}",
         "Antes de cada geração de geometria (v5.x → v5.x+1), executar o script\n"
         "FIORA_IC_geometry_v5_11.py — ele chama _check_clearances() automaticamente\n"
         "no início e levanta AssertionError se alguma folga vertical ficar < 50 mm.\n\n"
         "Bug histórico detectado (v5.9 → v5.10): cota nominal do separador era\n"
         "interpretada como pico do Λ, mas as placas inclinadas desciam 385 mm\n"
         "abaixo do pico, criando sobreposição com o topo dos eletrodos. Corrigido\n"
         "na v5.10 interpretando pct × H como BORDA INFERIOR das placas externas.")
ws12.merge_cells(f"A{r}:E{r}")
ws12[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws12.row_dimensions[r].height = 130

# ═══════════════════════════════════════════════════════
# ABA 13 — NOVA: Manifold de 4 ramais DN50
# ═══════════════════════════════════════════════════════
ws13 = wb.create_sheet("13_Manifold_Recirculacao")
for cl, w in [("A", 38), ("B", 18), ("C", 16), ("D", 49)]:
    ws13.column_dimensions[cl].width = w

set_cell(ws13, "B1", "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v4.0",
         bg=C_TITLE, font=FONT_WHITE_BOLD)
ws13.merge_cells("B1:D1")
set_cell(ws13, "B2", "ABA 13 — MANIFOLD DE 4 RAMAIS DN50 (★ NOVA v4.0)",
         bg=C_SUBTITLE, font=FONT_WHITE_BOLD)
ws13.merge_cells("B2:D2")

r = 4
section_header(ws13, r, "PARÂMETROS GEOMÉTRICOS — TUBO PRINCIPAL DN100")
r += 1
write_row(ws13, r, "Diâmetro nominal", 0.100, "m", "DN100 — Vinícius"); r += 1
write_row(ws13, r, "Espessura de parede", 0.004, "m", "4 mm — padrão DN100"); r += 1
write_row(ws13, r, "Diâmetro interno", "=B5-2*B6", "m", "= 92 mm"); r += 1
write_row(ws13, r, "Cota inferior (manifold)", 0.15, "m",
          "★ Vinícius 18/05 17:06: 'Até 15 cm do fundo'",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws13, r, "Cota superior (próximo ao domo)", 14.29, "m",
          "= H − TOP_DEPTH − 0,05"); r += 1
write_row(ws13, r, "Comprimento do tubo principal",
          "=B9-B8", "m", "= 14,14 m"); r += 1

r += 1
section_header(ws13, r, "PARÂMETROS GEOMÉTRICOS — 4 RAMAIS DN50")
r += 1
write_row(ws13, r, "Diâmetro nominal de cada ramal", 0.050, "m",
          "★ Vinícius 18/05 17:06: DN50 confirmado",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws13, r, "Espessura de parede", 0.004, "m", "4 mm"); r += 1
write_row(ws13, r, "Comprimento de cada ramal", 0.70, "m",
          "Vinícius: 'pode manter como propôs'",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws13, r, "Cota dos ramais (= manifold)", "=B8", "m",
          "Mesma cota do manifold"); r += 1
write_row(ws13, r, "Azimutes dos 4 ramais", "45° / 135° / 225° / 315°", "—",
          "Entre bocais, fora dos eletrodos",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws13, r, "Posição final do ramal (x,y) — az 45°",
          "(+0,495; +0,495)", "m", "= 0,7 × cos/sin(45°)"); r += 1
write_row(ws13, r, "Folga radial à parede", "=R_REATOR − 0,7",
          "m", "≈ 0,345 m"); r += 1
write_row(ws13, r, "Final dos ramais", "Aberto (descarga livre)", "—",
          "Vinícius: 'pode manter como propôs'"); r += 1

r += 1
section_header(ws13, r, "CONSERVAÇÃO DE ÁREA HIDRÁULICA", color=C_CALC_HD)
r += 1
write_row(ws13, r, "Área da seção do tubo principal (DN100)",
          "=PI()*(B5/2)^2", "m²", "= 7,854 ×10⁻³ m²"); r += 1
write_row(ws13, r, "Área de cada ramal (DN50)",
          "=PI()*(B13/2)^2", "m²", "= 1,963 ×10⁻³ m²"); r += 1
write_row(ws13, r, "Área total dos 4 ramais",
          f"=4*B{r-1}", "m²", "= 7,854 ×10⁻³ m²"); r += 1
write_row(ws13, r, "★ Razão Área_4ramais / Área_DN100",
          f"=B{r-1}/B{r-3}", "—", "= 1,000 ✓ Conservação OK"); r += 1
ws13[f"B{r-1}"].fill = fill(C_FIXED)

r += 1
section_header(ws13, r, "BC NO STAR-CCM+", color=C_NOTE_HD)
r += 1
set_cell(ws13, f"A{r}",
         "1. Toda a estrutura (tubo principal + 4 ramais) é UM ÚNICO baffle\n"
         "   No-slip Wall no STAR-CCM+ (separa fluxo interno descendente do externo ascendente).\n\n"
         "2. APENAS a face SUPERIOR do tubo principal é Velocity Inlet:\n"
         "   • Q = 350 m³/d (recirculação líquida)\n"
         "   • Direção: −Z (descendente)\n"
         "   • Temperatura: 35°C\n"
         "   • Fase: líquida\n\n"
         "3. Os 4 ramais descarregam LIVREMENTE no domínio (extremidades abertas).\n"
         "   No CFD, basta o sólido baffle ter aberturas nas extremidades.")
ws13.merge_cells(f"A{r}:D{r}")
ws13[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws13.row_dimensions[r].height = 175

# ═══════════════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════════════
wb.save(OUT)
print(f"✅ Salvo: {OUT}")
print(f"   Abas: {wb.sheetnames}")
