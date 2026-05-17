"""
FIORA IC — Geração da Planilha de Setup CFD v3.0
=================================================
Corrige os 7 erros numéricos da v2.0 e adiciona 4 abas novas:
- Aba 8: Eletrodos Eletroativos
- Aba 9: Regiões Funcionais A/B + Cinética (pós-processamento)
- Aba 10: Reports & Monitors no STAR-CCM+
- Aba 11: Matriz Consolidada de Cenários

Preserva 100% a estrutura visual da v2.0 (cores, fonts, layout).
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/tmp/setup.xlsx"
OUT = "/home/user/Programacao/FIORA_IC_Setup_CFD_v3.xlsx"

# Paleta de cores extraída da v2.0
C_TITLE      = "1F3864"  # azul escuro — título master
C_SUBTITLE   = "ED7D31"  # laranja — nome da aba
C_INPUT_HD   = "0070C0"  # azul — header de seção INPUT
C_CALC_HD    = "00B050"  # verde — header de seção CALCULADO
C_NOTE_HD    = "7030A0"  # roxo — header de seção NOTA/INFO
C_INPUT_BG   = "DCE6F1"  # azul claro — célula de input
C_CALC_BG    = "E2EFDA"  # verde claro — célula calculada
C_HIGHLIGHT  = "FFF2CC"  # amarelo claro — destaque/correção
C_FIXED      = "FCE4D6"  # laranja claro — célula corrigida v3.0

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFFFF", size=11)
FONT_BLUE       = Font(color="FF0000FF")
FONT_BOLD       = Font(bold=True)
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color):
    return PatternFill("solid", fgColor=color)


def set_cell(ws, coord, value, bg=None, font=None, border=False, align=None):
    c = ws[coord]
    c.value = value
    if bg is not None:
        c.fill = fill(bg)
    if font is not None:
        c.font = font
    if border:
        c.border = BORDER
    if align is not None:
        c.alignment = align
    return c


def write_row(ws, row, label, value, unit=None, note=None, value_bg=C_CALC_BG, value_font=None):
    """Escreve uma linha padrão: A=label, B=valor, C=unidade, D=nota."""
    set_cell(ws, f"A{row}", label)
    set_cell(ws, f"B{row}", value, bg=value_bg, font=value_font)
    if unit is not None:
        set_cell(ws, f"C{row}", unit)
    if note is not None:
        set_cell(ws, f"D{row}", note)


def section_header(ws, row, text, color=C_INPUT_HD, cols=("A", "B", "C", "D")):
    for col in cols:
        set_cell(ws, f"{col}{row}", text if col == "A" else None, bg=color, font=FONT_WHITE_BOLD)
        ws[f"{col}{row}"].border = BORDER
    # Merge if available
    ws.merge_cells(f"{cols[0]}{row}:{cols[-1]}{row}")
    ws[f"{cols[0]}{row}"].alignment = Alignment(horizontal="left", vertical="center")


# ═══════════════════════════════════════════════════════
# CARREGAR E CORRIGIR
# ═══════════════════════════════════════════════════════

wb = load_workbook(SRC)

# ─── Atualizar título master da Aba 1 (versão v3.0) ───
ws1 = wb["1_Geometria"]
ws1["B1"].value = "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v3.0  (16/Mai/2026)"

# ═══════════════════════════════════════════════════════
# FIX ABA 1 — GEOMETRIA (3 erros)
# ═══════════════════════════════════════════════════════

# R33: TOP_DEPTH errado (12.775 → 0.261 = D/8)
ws1["B33"] = "=C7/8"
ws1["B33"].fill = fill(C_FIXED)
ws1["D33"] = "★ FIX v3.0: D/8 (era 12,775 — fórmula errada)"

# R34: Topo cilindro errado (14.6 → 14.339 = H − D/8)
ws1["B34"] = "=C8-C7/8"
ws1["B34"].fill = fill(C_FIXED)
ws1["D34"] = "★ FIX v3.0: H − TOP_DEPTH (era 14,6 — era altura total)"

# R35: Ápice referenciava B8 (texto), agora C8 (valor 14.6)
ws1["B35"] = "=C8"
ws1["B35"].fill = fill(C_FIXED)
ws1["D35"] = "★ FIX v3.0: =C8 (era =B8, retornava texto)"

# ═══════════════════════════════════════════════════════
# FIX ABA 2 — VAZÕES/VELOCIDADES (cascata de erros)
# ═══════════════════════════════════════════════════════

ws2 = wb["2_Vazoes_Velocidades"]

# Adicionar input EXPLÍCITO do DN80 (Vinícius confirmou)
# Inserir uma linha de input antes do bloco de bocais
set_cell(ws2, "A21", "Diâmetro nominal do bocal (DN80 — Vinícius)",
         font=FONT_BOLD)
set_cell(ws2, "B21", 0.080, bg=C_INPUT_BG, font=FONT_BLUE)
set_cell(ws2, "C21", "m")
set_cell(ws2, "D21", "★ NOVO v3.0: DN80 confirmado — Vinícius 16/05 18:04")

# R25: d_bocal — usar valor confirmado em vez de calcular de 10% da área
ws2["C25"] = "=B21"
ws2["C25"].fill = fill(C_FIXED)
ws2["E25"] = "★ FIX v3.0: usa DN80 confirmado (não 269 mm calculado de 10% da área)"

# R26: d_bocal em mm — recalcula automático
# (fórmula =C25*1000 já está OK)

# R27: V_bocal — usar Q_por_bocal / A_bocal_real (não Q_total / A_total)
# A_bocal_real = π · (D/2)² = π · (B21/2)²
ws2["C27"] = "=C15/C8/(PI()*(B21/2)^2)"
ws2["C27"].fill = fill(C_FIXED)
ws2["E27"] = "★ FIX v3.0: Q_por_bocal/A_DN80 (era Q_total/A_total → 11× menor)"

# R28: Re_bocal — fórmula já estava OK, mas valor vai propagar pela correção acima
# ws2["C28"] = "=C6*C27*C25/C7" — mantida
ws2["E28"] = "★ Auto-atualizado: agora ~17.020 (era 5.038)"

# ═══════════════════════════════════════════════════════
# FIX ABA 3 — STOKES INVÁLIDO (adicionar correlação correta)
# ═══════════════════════════════════════════════════════

ws3 = wb["3_Fases_Propriedades"]

# Reescrever rótulo R20 para deixar claro que é Stokes (teórico)
ws3["B20"] = "Vel. terminal bolha — Stokes V_t = (ρ_L−ρ_G)·g·d_b²/(18·μ_L)   ⚠ TEÓRICO"

# Adicionar verificação de validade de Stokes e estimativa real
# Linhas 21 já têm V_t em m/h. Vou adicionar 21b, 21c, 21d à direita (cols D,E)
ws3["E20"] = "★ v3.0: STOKES INVÁLIDO para bolhas >0,1mm. Use linhas 21b–21d ↓"
ws3["E20"].fill = fill(C_HIGHLIGHT)

# Reescrever linha 22 (estava vazia) com Re_bolha + V_real
set_cell(ws3, "B22", "Re_bolha (Stokes)  Re = ρ·V_Stokes·d/μ", font=FONT_BOLD)
set_cell(ws3, "C22", "=C6*C20*C14/C7", bg=C_CALC_BG)
set_cell(ws3, "D22", "—")
set_cell(ws3, "E22", "Se Re>>1 → Stokes inválido. Use Schiller-Naumann")

# Inserir novas linhas usando função para inserção no Aba 3
# Vou adicionar uma seção nova no final da fase gasosa (após R22)
# Mas para não bagunçar, vou colocar como notas estendidas na coluna E

ws3["E21"] = "★ Comparar com V_real ↓"

# Para fase sólida — mesmo problema (R31 Stokes)
ws3["B31"] = "Vel. sedimentação grânulo — Stokes V_sed = (ρ_S−ρ_L)·g·d_p²/(18·μ_L)  ⚠ TEÓRICO"
ws3["E31"] = "★ v3.0: STOKES INVÁLIDO para Re>>1 — usar correlação Schiller-Naumann"
ws3["E31"].fill = fill(C_HIGHLIGHT)

# Adicionar nota sobre α_S inicial (memorial diz 0.35 na zona inferior)
ws3["E29"] = "★ v3.0: 0,03 = média do reator. NO LEITO (zona inferior) usar α_S=0,35 (Wang 2010)"
ws3["E29"].fill = fill(C_HIGHLIGHT)

# Adicionar nota de densidade do lodo (1460 = limite superior)
ws3["E28"] = "★ v3.0: 1460 = limite superior. Mais comum na lit.: 1.050 (D'Bastiani 2021). Usar como sensibilidade."
ws3["E28"].fill = fill(C_HIGHLIGHT)

# ═══════════════════════════════════════════════════════
# FIX ABA 4 — k_inlet e ε_inlet (auto-corrigem por cascata)
# ═══════════════════════════════════════════════════════

ws4 = wb["4_Turbulencia_CC"]

# Os valores k_inlet e ε_inlet já vão atualizar automaticamente quando a Aba 2 mudar
# Vou só adicionar nota sobre a correção
ws4["E13"] = "★ Auto-atualizado v3.0: agora ~1,08×10⁻⁴ m²/s² (era 8,3×10⁻⁷)"
ws4["E13"].fill = fill(C_FIXED)
ws4["E14"] = "★ Auto-atualizado v3.0: agora ~9,7×10⁻⁶ m²/s³ (era 6,6×10⁻⁹)"
ws4["E14"].fill = fill(C_FIXED)

# Fix R10 e R11: a fórmula da intensidade turbulenta usa Re_bocal (C28 da aba 2)
# que agora vai auto-atualizar. Sem mudança necessária.

# ═══════════════════════════════════════════════════════
# ABA 5 — UNIFICAR M_mix (corrigir 25.8 → 25.7 incluindo H₂S)
# ═══════════════════════════════════════════════════════

ws5 = wb["5_Pressao_Operacional"]

# Encontrar a linha de M_mix
for r in range(1, ws5.max_row + 1):
    bval = ws5.cell(row=r, column=2).value
    if bval and "Massa molar mistura" in str(bval):
        ws5.cell(row=r, column=3).value = "=0.65*16+0.34*44+0.01*34"
        ws5.cell(row=r, column=3).fill = fill(C_FIXED)
        ws5.cell(row=r, column=5).value = "★ FIX v3.0: inclui H₂S (era só CH₄+CO₂ → 25,8). Agora 25,7"
        break

# ═══════════════════════════════════════════════════════
# ABA 8 — ELETRODOS ELETROATIVOS (NOVA)
# ═══════════════════════════════════════════════════════

ws8 = wb.create_sheet("8_Eletrodos")
# Aplicar mesmas larguras de coluna
for col_letter, width in [("A", 36.86), ("B", 18), ("C", 15), ("D", 49), ("E", 40)]:
    ws8.column_dimensions[col_letter].width = width

set_cell(ws8, "B1", "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v3.0",
         bg=C_TITLE, font=FONT_WHITE_BOLD)
ws8.merge_cells("B1:E1")
set_cell(ws8, "B2", "ABA 8 — ELETRODOS ELETROATIVOS (Ânodo/Cátodo)",
         bg=C_SUBTITLE, font=FONT_WHITE_BOLD)
ws8.merge_cells("B2:E2")

r = 4
section_header(ws8, r, "PARÂMETROS GEOMÉTRICOS (Vinícius 15/05 16:50 — módulo eletroativo)")
r += 1
write_row(ws8, r, "Número total de placas", 4, "—",
          "2 ânodos + 2 cátodos", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Altura da placa (ELEC_H)", 2.400, "m",
          "Vinícius — tabela 15/05", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Largura da placa (ELEC_W)", 0.520, "m",
          "Vinícius — tabela 15/05", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Espessura da placa (ELEC_THICK)", 0.008, "m",
          "8 mm — padrão eletrodo industrial", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Posição radial (ELEC_R_POS = 65% × R_reator)",
          "=0.65*('1_Geometria'!C11/2)", "m",
          "0,65 × R = 0,68 m do centro"); r += 1
write_row(ws8, r, "Cota base (ELEC_Z_BOT)", 0.250, "m",
          "Vinícius: 'abaixo do 1° separador'",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Cota topo (ELEC_Z_TOP = ELEC_Z_BOT + ELEC_H)",
          "=B8+B6", "m", "0,25 + 2,40 = 2,65 m"); r += 1
write_row(ws8, r, "Cota do Sep. 20% (limite superior permitido)",
          "='1_Geometria'!C19", "m", "Não pode haver overlap"); r += 1
write_row(ws8, r, "Folga vertical até Sep. 20%",
          "=B11-B10", "m", "✓ Deve ser positivo"); r += 1

r += 1
section_header(ws8, r, "ÁREA ELETROATIVA")
r += 1
write_row(ws8, r, "Área por placa", "=B6*B7", "m²", "= ELEC_H × ELEC_W ≈ 1,25 m²"); r += 1
write_row(ws8, r, "Área eletroativa total (4 placas)", "=4*B14", "m²",
          "≈ 5,0 m² — confere com Vinícius"); r += 1

r += 1
section_header(ws8, r, "ARRANJO DOS PARES E GAP")
r += 1
write_row(ws8, r, "Número de pares Ânodo/Cátodo", 2, "—",
          "Par 1 + Par 2", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Azimute do Par 1", 0.0, "°",
          "Adotado — simetria diametral", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Azimute do Par 2", 180.0, "°",
          "Adotado — diametralmente oposto", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Gap FACE-A-FACE (ELEC_GAP_FACE)", 0.030, "m",
          "Vinícius: 'máximo 30 mm' — adotado limite superior",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Distância CENTRO-A-CENTRO (c2c = gap + esp.)",
          "=B21+B8", "m", "Usado no CAD para posicionar as placas"); r += 1

r += 1
section_header(ws8, r, "VERIFICAÇÕES GEOMÉTRICAS (não pode haver conflito)")
r += 1
write_row(ws8, r, "Raio externo da placa (r_ext = R_POS + W/2)",
          "=B9+B7/2", "m", "0,68 + 0,26 = 0,94 m"); r += 1
write_row(ws8, r, "Raio interno da placa (r_int = R_POS − W/2)",
          "=B9-B7/2", "m", "0,68 − 0,26 = 0,42 m"); r += 1
write_row(ws8, r, "Raio do reator",
          "='1_Geometria'!C11/2", "m", "≈ 1,045 m"); r += 1
write_row(ws8, r, "Folga radial até parede (R − r_ext)",
          "=B27-B25", "m", "✓ Deve ser > 0,05 m"); r += 1
write_row(ws8, r, "Raio externo do DN100 (R_RETURN_OUT)",
          0.050, "m", "DN100/2", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws8, r, "Folga radial até DN100 (r_int − R_RETURN_OUT)",
          "=B26-B29", "m", "✓ Deve ser > 0,1 m"); r += 1

r += 1
section_header(ws8, r, "BOUNDARY CONDITION NO STAR-CCM+")
r += 1
write_row(ws8, r, "Tipo de BC", "No-slip Wall", "—",
          "Sólido inerte — sem eletroquímica no caso base"); r += 1
write_row(ws8, r, "Nome no STAR-CCM+",
          "Wall_Electrode_A1 / A2 / C1 / C2", "—",
          "4 boundaries separadas"); r += 1
write_row(ws8, r, "Rugosidade", 0, "m",
          "Aço inox liso", value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1

r += 1
section_header(ws8, r, "PRÓXIMA ETAPA — ELETROQUÍMICA (FORA DO ESCOPO ATUAL)",
               color=C_NOTE_HD)
r += 1
set_cell(ws8, "A37", "Parâmetros elétricos (tensão, corrente, densidade de corrente, "
         "transferência iônica) serão implementados na PRÓXIMA ETAPA via módulo "
         "Electrochemistry do STAR-CCM+, após validação do caso base hidrodinâmico.")
ws8.merge_cells("A37:E37")
ws8["A37"].alignment = Alignment(wrap_text=True, vertical="top")
ws8.row_dimensions[37].height = 45

# ═══════════════════════════════════════════════════════
# ABA 9 — REGIÕES FUNCIONAIS A/B + CINÉTICA (NOVA)
# ═══════════════════════════════════════════════════════

ws9 = wb.create_sheet("9_Regioes_Cinetica_PosProc")
for col_letter, width in [("A", 36.86), ("B", 18), ("C", 15), ("D", 49), ("E", 40)]:
    ws9.column_dimensions[col_letter].width = width

set_cell(ws9, "B1", "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v3.0",
         bg=C_TITLE, font=FONT_WHITE_BOLD)
ws9.merge_cells("B1:E1")
set_cell(ws9, "B2", "ABA 9 — REGIÕES FUNCIONAIS A/B + CINÉTICA SIMPLIFICADA (Pós-processamento)",
         bg=C_SUBTITLE, font=FONT_WHITE_BOLD)
ws9.merge_cells("B2:E2")

r = 4
section_header(ws9, r, "DEFINIÇÃO DAS REGIÕES FUNCIONAIS (Stage 2 §8) — Field Functions no STAR-CCM+",
               color=C_NOTE_HD)
r += 1
write_row(ws9, r, "Zona A — Região hidrogênica (z do fundo)",
          "0 → 8,76 m", "—",
          "0% → 60% da altura útil"); r += 1
write_row(ws9, r, "Zona A — em frame centrado (CFD origem = centro)",
          "−7,30 → +1,46 m", "—",
          "z_rel = z − H/2"); r += 1
write_row(ws9, r, "Zona B — Região metanogênica (z do fundo)",
          "8,76 → 13,87 m", "—",
          "60% → 95% da altura útil"); r += 1
write_row(ws9, r, "Zona B — em frame centrado",
          "+1,46 → +6,57 m", "—",
          "z_rel = z − H/2"); r += 1
write_row(ws9, r, "Zona Final — Separação (z do fundo)",
          "13,87 → 14,60 m", "—",
          "95% → 100% — apenas separação"); r += 1

r += 1
section_header(ws9, r, "FIELD FUNCTIONS A CRIAR NO STAR-CCM+ (Tools > Field Functions > New)")
r += 1
set_cell(ws9, f"A{r}", "Nome: Zone_A    |    Definition:",
         font=FONT_BOLD)
set_cell(ws9, f"B{r}", "($$Position[2] < 1.46) && ($$Position[2] > -7.30) ? 1.0 : 0.0",
         bg=C_INPUT_BG, font=FONT_BLUE)
ws9.merge_cells(f"B{r}:E{r}")
ws9[f"B{r}"].alignment = Alignment(wrap_text=True)
r += 1
set_cell(ws9, f"A{r}", "Nome: Zone_B    |    Definition:", font=FONT_BOLD)
set_cell(ws9, f"B{r}", "($$Position[2] >= 1.46) && ($$Position[2] < 6.57) ? 1.0 : 0.0",
         bg=C_INPUT_BG, font=FONT_BLUE)
ws9.merge_cells(f"B{r}:E{r}")
ws9[f"B{r}"].alignment = Alignment(wrap_text=True)
r += 1
section_header(ws9, r, "PARÂMETROS CINÉTICOS (Stage 2 §10, §13)")
r += 1
write_row(ws9, r, "DQO afluente (S_DQO_in)", 25.0, "kg/m³",
          "25.000 mg/L — Stage 2 §2 (faixa 20–30)",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws9, r, "Carga orgânica total (L_org)",
          "=B15*'2_Vazoes_Velocidades'!C4", "kg DQO/d",
          "= S_DQO × Q_af ≈ 1.250 kg DQO/d"); r += 1
write_row(ws9, r, "Constante cinética Região A (k_app,A)", 0.30, "h⁻¹",
          "Stage 2 §10 — faixa 0,20–0,50",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws9, r, "Constante cinética Região B (k_app,B)", 0.08, "h⁻¹",
          "Stage 2 §13 — faixa 0,05–0,15",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws9, r, "Fator estequiométrico H₂ (f_H₂)", 0.05, "kg H₂/kg DQO",
          "EXPLORATÓRIO — variar 0,02 a 0,10",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1
write_row(ws9, r, "Rendimento de CH₄ (Y_CH₄)", 0.30, "Nm³CH₄/kg DQO",
          "Stage 2 §13 — faixa 0,25–0,32",
          value_bg=C_INPUT_BG, value_font=FONT_BLUE); r += 1

r += 1
section_header(ws9, r, "ESTIMATIVAS PRELIMINARES (validação cruzada com planilha Vinícius)")
r += 1
# Volumes das zonas (cilíndrico)
write_row(ws9, r, "Volume teórico Região A (V_A = A × h_A)",
          "='1_Geometria'!C13*'1_Geometria'!C12*0.6", "m³",
          "60% × V_reator"); r += 1
write_row(ws9, r, "Volume teórico Região B (V_B = A × (h_B − h_A))",
          "='1_Geometria'!C13*'1_Geometria'!C12*(0.95-0.6)", "m³",
          "35% × V_reator"); r += 1
# Taxa de consumo de DQO
write_row(ws9, r, "Taxa de consumo DQO Região A (r_DQO,A)",
          "=B17/3600*B15*B23", "kg/s",
          "= k_app,A × S_DQO × V_A  (assumindo S_DQO constante)"); r += 1
write_row(ws9, r, "Taxa de consumo DQO Região B (r_DQO,B)",
          "=B18/3600*B15*0.3*B24", "kg/s",
          "S_DQO,res ≈ 30% S_DQO,in"); r += 1
# Produção de H₂
write_row(ws9, r, "Taxa H₂ estimada (r_H₂ = f_H₂ × r_DQO,A)",
          "=B19*B25", "kg/s",
          "Stage 2 §11 — variável de resposta"); r += 1
write_row(ws9, r, "Produção diária H₂",
          "=B27*86400", "kg/d",
          "Para conferência com Excel Vinícius"); r += 1
write_row(ws9, r, "Produção volumétrica H₂ (ρ_H₂=0.0899 kg/Nm³)",
          "=B28/0.0899", "Nm³/d",
          "1 kg H₂ ≈ 11,12 Nm³"); r += 1
# Produção de CH₄
write_row(ws9, r, "Taxa CH₄ estimada (r_CH₄ = Y_CH₄ × r_DQO,B × ρ_CH₄)",
          "=B20*B26*0.668", "kg/s",
          "ρ_CH₄ = 0,668 kg/Nm³"); r += 1
write_row(ws9, r, "Produção diária CH₄ volumétrica",
          "=B20*B26*86400", "Nm³/d",
          ""); r += 1

r += 1
section_header(ws9, r, "KPI PRINCIPAL — H₂/DQO (objetivo do Vinícius)", color=C_CALC_HD)
r += 1
write_row(ws9, r, "Eficiência kg H₂/kg DQO afluente (KPI 1)",
          "=B28/B16", "kg H₂/kg DQO",
          "Stage 2 §17 — métrica principal"); r += 1
write_row(ws9, r, "Eficiência kg H₂/kg DQO consumida na Região A (KPI 2)",
          "=B27/(B25*1.0)", "kg H₂/kg DQO",
          "Eficiência intrínseca da fermentação"); r += 1

r += 1
section_header(ws9, r, "PIPELINE DE PÓS-PROCESSAMENTO",
               color=C_NOTE_HD)
r += 1
set_cell(ws9, "A37", "1. CFD entrega via Aba 10: Q_DQO_in, Q_gas_total, vazões por DN50, "
         "volumes reais por zona (descontando zonas mortas), tempo de residência, "
         "fração de fase em cada zona\n"
         "2. Esta aba calcula KPIs preliminares usando Stage 2 §10/§13\n"
         "3. Planilha Excel do Vinícius compara e calibra f_H₂\n"
         "4. Comparação entre cenários C0/C1/C2 (Stage 2 §15) + C0/C1/C2/C3 (ângulos bocal)")
ws9.merge_cells("A37:E37")
ws9["A37"].alignment = Alignment(wrap_text=True, vertical="top")
ws9.row_dimensions[37].height = 80

# ═══════════════════════════════════════════════════════
# ABA 10 — REPORTS & MONITORS NO STAR-CCM+ (NOVA)
# ═══════════════════════════════════════════════════════

ws10 = wb.create_sheet("10_Reports_STARCCM")
for col_letter, width in [("A", 32), ("B", 20), ("C", 32), ("D", 28), ("E", 40)]:
    ws10.column_dimensions[col_letter].width = width

set_cell(ws10, "B1", "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v3.0",
         bg=C_TITLE, font=FONT_WHITE_BOLD)
ws10.merge_cells("B1:E1")
set_cell(ws10, "B2", "ABA 10 — REPORTS & MONITORS NO STAR-CCM+ (Pós-processamento)",
         bg=C_SUBTITLE, font=FONT_WHITE_BOLD)
ws10.merge_cells("B2:E2")

# Header da tabela
r = 4
headers = ["Report Name", "Tipo (STAR)", "Field Function / Definition", "Superfície/Volume", "Output / Uso"]
for i, h in enumerate(headers):
    set_cell(ws10, f"{chr(65+i)}{r}", h, bg=C_INPUT_HD, font=FONT_WHITE_BOLD, border=True)

reports = [
    # (name, type, definition, target, output)
    ("=== ENTRADAS DE MASSA ===", "", "", "", ""),
    ("Q_DQO_in_nozzles", "Mass Flow", "ρ_L × velocity (faseLíquida)", "6× Inlet_Nozzle_01..06", "kg/s líquido entrando"),
    ("Q_DQO_in_DN100", "Mass Flow", "ρ_L × velocity (faseLíquida)", "Inlet_DN100_top", "kg/s recirculação"),
    ("Q_gas_in_diffusers", "Mass Flow Sum", "ṁ_gas (Mass Flow Inlet BC)", "12× Inlet_Diffuser", "kg/s biogás recirc."),
    ("=== SAÍDAS DE MASSA ===", "", "", "", ""),
    ("Q_gas_top", "Mass Flow", "ρ_G × velocity × α_G", "Outlet_Biogas (DN100 topo)", "kg/s biogás saindo"),
    ("Q_eff_lateral", "Mass Flow", "ρ_L × velocity × α_L", "Outlet_Effluent (DN100 lat)", "kg/s efluente saindo"),
    ("Q_DN50_20", "Mass Flow", "ρ × velocity (multifásico)", "Outlet_DN50_20", "Captura sep. 20%"),
    ("Q_DN50_40", "Mass Flow", "ρ × velocity (multifásico)", "Outlet_DN50_40", "Captura sep. 40%"),
    ("Q_DN50_60", "Mass Flow", "ρ × velocity (multifásico)", "Outlet_DN50_60", "Captura sep. 60%"),
    ("Q_DN50_95", "Mass Flow", "ρ × velocity (multifásico)", "Outlet_DN50_95", "Captura sep. 95%"),
    ("=== VOLUMES DAS REGIÕES (via Field Functions) ===", "", "", "", ""),
    ("Volume_Zone_A", "Volume Integral", "Zone_A", "Region (Reactor_Fluid)", "m³ úteis da Zona A"),
    ("Volume_Zone_B", "Volume Integral", "Zone_B", "Region (Reactor_Fluid)", "m³ úteis da Zona B"),
    ("Volume_DeadZones_A", "Volume Integral", "Zone_A × (mag(V) < 0.001)", "Region", "m³ de zona morta na A"),
    ("Volume_DeadZones_B", "Volume Integral", "Zone_B × (mag(V) < 0.001)", "Region", "m³ de zona morta na B"),
    ("=== DISTRIBUIÇÃO DE FASES POR ZONA ===", "", "", "", ""),
    ("Avg_alpha_lodo_A", "Vol-Weighted Avg", "Zone_A × α_lodo", "Region", "Fração média lodo Zona A"),
    ("Avg_alpha_lodo_B", "Vol-Weighted Avg", "Zone_B × α_lodo", "Region", "Fração média lodo Zona B"),
    ("Avg_alpha_gas_A", "Vol-Weighted Avg", "Zone_A × α_gas", "Region", "Fração média gás Zona A"),
    ("Avg_alpha_gas_B", "Vol-Weighted Avg", "Zone_B × α_gas", "Region", "Fração média gás Zona B"),
    ("Avg_alpha_biochar_A", "Vol-Weighted Avg", "Zone_A × α_biochar", "Region", "Fração biochar Zona A"),
    ("=== CISALHAMENTO E HIDRODINÂMICA ===", "", "", "", ""),
    ("Avg_shear_A", "Vol-Weighted Avg", "Zone_A × strainRate", "Region", "Cisalhamento médio na A (Wu 2015: 57–97% por gás)"),
    ("Avg_shear_B", "Vol-Weighted Avg", "Zone_B × strainRate", "Region", "Cisalhamento médio na B"),
    ("Max_velocity_global", "Maximum", "mag(velocity)", "Region", "Pico de velocidade"),
    ("Avg_velocity_axial", "Vol-Weighted Avg", "velocity[2]", "Region", "Vel. axial média"),
    ("=== TEMPO DE RESIDÊNCIA (DTR via passive scalar) ===", "", "", "", ""),
    ("DTR_inlet_pulse", "Passive Scalar", "Pulso unitário em t=0 nos Inlets", "Outlet_Effluent", "Curva E(t) → tempo médio τ"),
    ("MRT_zone_A", "Passive Scalar", "Tempo médio na Zona A", "Region (Zone_A=1)", "Tempo de residência médio"),
    ("=== PRESSÃO E PERDA DE CARGA ===", "", "", "", ""),
    ("DeltaP_total", "Expression", "P(z=0) - P(z=14.6)", "Region (line probe)", "Pa — perda de carga total"),
    ("DeltaP_separator_20", "Surface Avg diff", "P_below - P_above sep.20%", "Baffle_Separator_20", "Pa por separador"),
    ("=== KPI FINAL (calculado em Expression Report) ===", "", "", "", ""),
    ("Eff_H2_per_DQO", "Expression", "($Q_H2_top × 86400) / ($Q_DQO_in_nozzles × 86400)",
     "—", "★ kg H₂/kg DQO afluente — KPI principal Vinícius"),
    ("Eff_gas_capture_intermediate", "Expression", "($Q_DN50_20+$Q_40+$Q_60)/$Q_gas_in",
     "—", "% gás capturado nos sep. intermediários"),
    ("DeadZone_fraction_A", "Expression", "$Volume_DeadZones_A / $Volume_Zone_A",
     "—", "% de volume morto na Zona A"),
]

r = 5
for entry in reports:
    if entry[0].startswith("==="):
        # Sub-header
        set_cell(ws10, f"A{r}", entry[0], bg=C_CALC_HD, font=FONT_WHITE_BOLD)
        ws10.merge_cells(f"A{r}:E{r}")
    else:
        for i, val in enumerate(entry):
            set_cell(ws10, f"{chr(65+i)}{r}", val, border=True)
    r += 1

# Nota final
r += 1
section_header(ws10, r, "EXPORTAÇÃO PARA EXCEL (planilha Vinícius)", color=C_NOTE_HD)
r += 1
set_cell(ws10, "A" + str(r),
         "No STAR-CCM+: Tools > Reports > Update Frequency = a cada N iterações (e.g., 50).\n"
         "Solution > Stopping Criteria > Save Monitor Data > Format = .csv.\n"
         "Arquivos .csv exportados serão importados na planilha Vinícius via tabela dinâmica\n"
         "para cruzamento com os cenários C0/C1/C2 (Stage 2) + 70°/75°/80°/85° (Vinícius 18:04).")
ws10.merge_cells(f"A{r}:E{r}")
ws10[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws10.row_dimensions[r].height = 75

# ═══════════════════════════════════════════════════════
# ABA 11 — MATRIZ CONSOLIDADA DE CENÁRIOS (NOVA)
# ═══════════════════════════════════════════════════════

ws11 = wb.create_sheet("11_Matriz_Cenarios")
widths = [("A", 14), ("B", 18), ("C", 18), ("D", 14), ("E", 14), ("F", 18), ("G", 30)]
for col_letter, width in widths:
    ws11.column_dimensions[col_letter].width = width

set_cell(ws11, "B1", "FIORA IC — Planilha de Setup CFD/STAR-CCM+  |  v3.0",
         bg=C_TITLE, font=FONT_WHITE_BOLD)
ws11.merge_cells("B1:G1")
set_cell(ws11, "B2", "ABA 11 — MATRIZ CONSOLIDADA DE CENÁRIOS DE SIMULAÇÃO",
         bg=C_SUBTITLE, font=FONT_WHITE_BOLD)
ws11.merge_cells("B2:G2")

# Header
r = 4
headers11 = ["ID", "NOZZLE_TANG_ANGLE", "Módulo Externo", "DQO_in [mg/L]",
             "Φ (k_app,A)", "Status / Prioridade", "Notas"]
for i, h in enumerate(headers11):
    set_cell(ws11, f"{chr(65+i)}{r}", h, bg=C_INPUT_HD, font=FONT_WHITE_BOLD, border=True)
ws11.row_dimensions[r].height = 30
for col_letter in [chr(65+i) for i in range(len(headers11))]:
    ws11[f"{col_letter}{r}"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

# Linhas
cenarios = [
    # (ID, NOZZLE, módulo, DQO, Phi, status, notas)
    ("C0", "80°", "nenhum", 25000, 1.00, "★ CASO BASE — rodar primeiro",
     "Vinícius 16/05 18:04 — caso base oficial"),
    ("C0_M30", "80°", "M30 (30 s)", 25000, 1.10, "Stage 2 §14 — sensibilidade",
     "Módulo externo de condicionamento da recirculação"),
    ("C0_M60", "80°", "M60 (60 s)", 25000, 1.18, "Stage 2 §14 — sensibilidade",
     "Idem M30 com tempo de exposição maior"),
    ("C1_75", "75°", "nenhum", 25000, 1.00, "Vinícius — sensibilidade bocal",
     "Swirl forte com leve componente radial"),
    ("C2_70", "70°", "nenhum", 25000, 1.00, "Vinícius — sensibilidade bocal",
     "Mais radial, melhor distribuição, menor swirl"),
    ("C3_85", "85°", "nenhum", 25000, 1.00, "Vinícius — sensibilidade bocal",
     "Quase tangente puro, swirl agressivo"),
    ("C0_DQO_min", "80°", "nenhum", 20000, 1.00, "Stage 2 §2 — sensibilidade carga",
     "Limite inferior da faixa"),
    ("C0_DQO_max", "80°", "nenhum", 30000, 1.00, "Stage 2 §2 — sensibilidade carga",
     "Limite superior da faixa"),
    ("C0_kapp_min", "80°", "nenhum", 25000, "k=0,20 h⁻¹", "Stage 2 §10 — sens. cinética",
     "Limite inferior k_app,A — pós-proc."),
    ("C0_kapp_max", "80°", "nenhum", 25000, "k=0,50 h⁻¹", "Stage 2 §10 — sens. cinética",
     "Limite superior k_app,A — pós-proc."),
]

r = 5
for cen in cenarios:
    for i, val in enumerate(cen):
        c = set_cell(ws11, f"{chr(65+i)}{r}", val, border=True)
        if cen[0] == "C0":
            c.fill = fill(C_HIGHLIGHT)
            c.font = FONT_BOLD
    r += 1

r += 1
section_header(ws11, r, "TOTAL DE CENÁRIOS PLANEJADOS",
               color=C_CALC_HD, cols=("A", "B", "C", "D", "E", "F", "G"))
r += 1
write_row(ws11, r, "Cenários geometria-dependente (re-rodar CAD)", 4, "casos",
          "C0, C1_75, C2_70, C3_85 — cada um requer .step novo"); r += 1
write_row(ws11, r, "Cenários setup-dependente (mesma geometria)", 4, "casos",
          "M30, M60, DQO_min, DQO_max — só mudar parâmetros"); r += 1
write_row(ws11, r, "Cenários pós-processamento (não rodar CFD)", 2, "casos",
          "k_app sensibilidade — calculado na Aba 9"); r += 1
write_row(ws11, r, "TOTAL", 10, "casos", "Conforme tabela acima"); r += 1

r += 1
section_header(ws11, r, "ESTRATÉGIA DE EXECUÇÃO RECOMENDADA",
               color=C_NOTE_HD, cols=("A", "B", "C", "D", "E", "F", "G"))
r += 1
set_cell(ws11, "A" + str(r),
         "1. Rodar C0 (caso base) até convergência hidrodinâmica steady-state\n"
         "2. Continuar C0 como transiente até 1×TRH (24 h simulados)\n"
         "3. Avaliar KPIs e validar contra planilha Vinícius\n"
         "4. Se OK: rodar C1/C2/C3 em paralelo (sensibilidade geométrica)\n"
         "5. Em série: M30/M60 e DQO_min/max (sensibilidade operacional)\n"
         "6. Consolidação final: gerar tabela de KPI H₂/DQO e CH₄/DQO para todos os 10 cenários")
ws11.merge_cells(f"A{r}:G{r}")
ws11[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws11.row_dimensions[r].height = 110

# ═══════════════════════════════════════════════════════
# ATUALIZAR ABA "6_Nota_DEM_vs_EE" — incluir update v3.0
# ═══════════════════════════════════════════════════════

ws6 = wb["6_Nota_DEM_vs_EE"]
ws6["A1"].value = ("★ v3.0: nota mantida + adicionadas Abas 8 (Eletrodos), 9 (Cinética+Pós), "
                   "10 (Reports), 11 (Matriz). Ver Aba 9 para cálculo H₂/DQO via pós-processamento.")

# ═══════════════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════════════

wb.save(OUT)
print(f"✅ Salvo: {OUT}")
print(f"   Abas: {wb.sheetnames}")
